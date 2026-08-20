"""Lectura del archivo de órdenes por ejecutar: validación, limpieza y filtrado.

Lo exporta el sistema operativo, no el ETL, así que no se puede dar nada por
sentado: encima de la cabecera viene un bloque de reporte de alto variable, el
texto llega con la codificación rota y nueve de cada diez filas todavía no
tienen técnico asignado.

Aquí solo se lee y se limpia; el cruce con el histórico es de otra capa. Los
nombres se traducen al vocabulario del tablero (`Localidad` -> barrio,
`Tip_Orden` -> tipo_os) para no arrastrar dos formas de llamar a lo mismo.
"""

from __future__ import annotations

import csv
import io
import logging
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.taxonomy import norm, norm_dato

logger = logging.getLogger(__name__)

# Cabecera del export -> campo propio. La clave está normalizada con `norm_dato`,
# así que aguanta tildes, mayúsculas, espacios y signos ('Tip_Orden', 'N° Medidor').
# Lo que no esté aquí se ignora: el archivo trae columnas vacías y otras que no
# usamos, y arrastrarlas solo abulta el cargue.
COLUMNAS: dict[str, str] = {
    "orden": "orden",
    "nic": "nic",
    "tiporden": "tipo_os",
    "tecnico": "tecnico",
    "numlote": "lote",
    "municipio": "municipio",
    "localidad": "barrio",
    "direccion": "direccion",
    "nombrecliente": "cliente",
    "tarifa": "tarifa",
    "deudavencida": "deuda",
    "facturasvencidas": "facturas",
    "antiguedad": "antiguedad",
    "estado": "estado",
    "tiposuspensionscr": "tipo_suspension",
    "estadodelservicio": "estado_servicio",
    "comentario": "comentario",
}

# Sin estas no hay cargue posible, así que sirven para dos cosas a la vez:
# localizar la fila de la cabecera y rechazar un Excel que no sea de órdenes.
OBLIGATORIAS: frozenset[str] = frozenset(
    {"orden", "nic", "tipo_os", "tecnico", "municipio", "barrio", "direccion", "estado"}
)

# La cabecera venía en la fila 5 del archivo de referencia, pero el bloque de
# reporte que la precede cambia de alto. Se busca en las primeras filas y no en
# todo el archivo: si no apareció aquí, no es un export de órdenes.
MAX_FILAS_CABECERA = 30

# Marca del sistema para una orden que todavía no tiene técnico.
SIN_ASIGNAR = norm_dato("No asignado")


class ArchivoInvalido(ValueError):
    """El archivo no se puede leer como un export de órdenes."""


@dataclass(frozen=True, slots=True)
class Orden:
    """Una orden por ejecutar, ya limpia.

    `bkey` se arma aquí ('MUNICIPIO | BARRIO') porque es la clave con la que el
    histórico identifica un barrio: dejarla lista evita que cada consumidor la
    vuelva a componer y se le olvide el formato.
    """

    orden: str
    nic: str
    tecnico: str
    tipo_os: str | None = None
    lote: str | None = None
    municipio: str | None = None
    barrio: str | None = None
    bkey: str | None = None
    direccion: str | None = None
    # Dato personal: no debe viajar hacia el modelo. Se conserva porque el
    # operativo sí necesita saber a nombre de quién está el suministro.
    cliente: str | None = None
    tarifa: str | None = None
    deuda: float | None = None
    facturas: int | None = None
    antiguedad: int | None = None
    estado: str | None = None
    tipo_suspension: str | None = None
    estado_servicio: str | None = None
    comentario: str | None = None


@dataclass(frozen=True, slots=True)
class Cargue:
    """Resultado de leer un archivo: lo que quedó y lo que se descartó.

    Los descartes se cuentan y se devuelven porque el filtro por técnico se lleva
    la mayor parte del archivo: si el número no se muestra, quien sube 10.000
    órdenes y ve 700 va a pensar que el cargue falló.
    """

    ordenes: list[Orden]
    fila_cabecera: int
    leidas: int
    descartadas: dict[str, int] = field(default_factory=dict)


def leer_ordenes(datos: bytes, nombre: str) -> Cargue:
    """Lee, valida y limpia un archivo de órdenes.

    Args:
        datos: contenido del archivo.
        nombre: nombre original; de él sale el formato.

    Returns:
        El cargue, con solo las órdenes que ya tienen técnico asignado.

    Raises:
        ArchivoInvalido: si el formato no se admite, el archivo está dañado o no
            tiene la cabecera de un export de órdenes.
    """
    extension = nombre.rsplit(".", 1)[-1].lower() if "." in nombre else ""
    if extension == "csv":
        filas = _filas_csv(datos)
    elif extension == "xlsx":
        filas = _filas_excel(datos)
    elif extension == "xls":
        raise ArchivoInvalido(
            "El formato .xls (Excel 97-2003) no se puede leer. Vuelve a guardar el "
            "archivo como .xlsx o .csv y súbelo de nuevo."
        )
    else:
        raise ArchivoInvalido(
            f"No se admiten archivos .{extension or 'sin extensión'}. Sube un .xlsx o un .csv."
        )

    return _construir(filas)


# --- Lectura cruda ------------------------------------------------------------


def _filas_excel(datos: bytes) -> list[list[Any]]:
    try:
        libro = load_workbook(io.BytesIO(datos), read_only=True, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError, OSError, ValueError) as exc:
        raise ArchivoInvalido(
            "No se pudo abrir el archivo. Comprueba que sea un Excel válido y que no esté dañado."
        ) from exc
    try:
        hoja = libro.worksheets[0]
        return [list(fila) for fila in hoja.iter_rows(values_only=True)]
    finally:
        libro.close()


def _filas_csv(datos: bytes) -> list[list[Any]]:
    """Decodifica y parte el CSV.

    El separador se detecta porque estos export salen con coma o con punto y
    coma según la configuración regional de quien los genera.
    """
    try:
        texto = datos.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = datos.decode("latin-1")

    muestra = texto[:4096]
    try:
        separador = csv.Sniffer().sniff(muestra, delimiters=",;\t|").delimiter
    except csv.Error:
        separador = ";" if muestra.count(";") > muestra.count(",") else ","

    return [list(fila) for fila in csv.reader(io.StringIO(texto), delimiter=separador)]


# --- Cabecera -----------------------------------------------------------------


def _buscar_cabecera(filas: list[list[Any]]) -> tuple[int, dict[int, str]]:
    """Devuelve (índice de la fila de cabecera, {columna -> campo}).

    Raises:
        ArchivoInvalido: si ninguna de las primeras filas trae las obligatorias.
    """
    mejor_fila, mejor_encontradas = -1, set()

    for i, fila in enumerate(filas[:MAX_FILAS_CABECERA]):
        mapa = {
            j: COLUMNAS[clave]
            for j, celda in enumerate(fila)
            if (clave := norm_dato(celda)) in COLUMNAS
        }
        encontradas = set(mapa.values()) & OBLIGATORIAS
        if encontradas == OBLIGATORIAS:
            return i, mapa
        if len(encontradas) > len(mejor_encontradas):
            mejor_fila, mejor_encontradas = i, encontradas

    detalle = (
        f"La fila {mejor_fila + 1} se le parece, pero le faltan estas columnas: "
        f"{', '.join(sorted(OBLIGATORIAS - mejor_encontradas))}."
        if mejor_fila >= 0
        else "No se encontró ninguna fila de cabecera."
    )
    raise ArchivoInvalido(
        f"El archivo no parece un export de órdenes. {detalle} "
    )


# --- Limpieza de valores ------------------------------------------------------


def _texto(valor: Any) -> str | None:
    """Texto limpio, o None si no hay dato.

    Repara de paso la doble codificación ('SuspensiÃ³n' -> 'Suspensión'): el
    export la trae en las columnas de texto libre y, sin arreglarla, lo que el
    chat acabe citando queda ilegible.
    """
    if valor is None:
        return None
    texto = str(valor)
    if "Ã" in texto or "Â" in texto:
        try:
            texto = texto.encode("latin-1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
    return " ".join(texto.split()) or None


def _comentario(valor: Any) -> str | None:
    """Igual que `_texto`, pero sin el punto final, que el export pone o no."""
    texto = _texto(valor)
    return (texto.rstrip(" .") or None) if texto else None


def _identificador(valor: Any) -> str | None:
    """Orden y NIC son etiquetas, no cantidades: nunca deben quedar como '1.0'."""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return _texto(valor)


def _decimal(valor: Any) -> float | None:
    if valor is None or valor == "":
        return None
    try:
        return float(str(valor).replace(",", "."))
    except ValueError:
        return None


def _entero(valor: Any) -> int | None:
    numero = _decimal(valor)
    return int(numero) if numero is not None else None


def _canonizar_comentarios(registros: list[dict[str, Any]]) -> None:
    """Unifica las variantes de un mismo comentario.

    El sistema exporta el mismo texto con tildes y sin ellas ('Suspensión del
    servicio de energía' y 'Suspension del servicio de energia'), así que contar
    por el texto crudo parte un grupo en dos. Se agrupan por su forma sin tildes
    y gana la variante más repetida; a igual frecuencia, la que las conserva.
    """
    variantes: dict[str, Counter[str]] = defaultdict(Counter)
    for registro in registros:
        if comentario := registro["comentario"]:
            variantes[norm(comentario)][comentario] += 1

    canon = {
        clave: max(cuenta.items(), key=lambda par: (par[1], _acentos(par[0])))[0]
        for clave, cuenta in variantes.items()
    }
    for registro in registros:
        if comentario := registro["comentario"]:
            registro["comentario"] = canon[norm(comentario)]


def _acentos(texto: str) -> int:
    return sum(1 for caracter in texto if not caracter.isascii())


# --- Armado -------------------------------------------------------------------


def _construir(filas: list[list[Any]]) -> Cargue:
    inicio, mapa = _buscar_cabecera(filas)

    registros: list[dict[str, Any]] = []
    descartadas = {"sin_orden": 0, "sin_tecnico": 0, "duplicadas": 0}
    vistas: set[str] = set()
    leidas = 0

    for fila in filas[inicio + 1 :]:
        crudo = {campo: fila[j] if j < len(fila) else None for j, campo in mapa.items()}
        if all(valor is None or str(valor).strip() == "" for valor in crudo.values()):
            continue  # filas en blanco y pies del reporte, que no son descartes
        leidas += 1

        orden = _identificador(crudo["orden"])
        if not orden:
            descartadas["sin_orden"] += 1
            continue

        tecnico = _texto(crudo["tecnico"])
        if not tecnico or norm_dato(tecnico) == SIN_ASIGNAR:
            descartadas["sin_tecnico"] += 1
            continue

        if orden in vistas:
            descartadas["duplicadas"] += 1
            continue
        vistas.add(orden)

        municipio = _texto(crudo["municipio"])
        barrio = _texto(crudo["barrio"])
        registros.append(
            {
                "orden": orden,
                "nic": _identificador(crudo["nic"]) or "",
                "tecnico": tecnico,
                "tipo_os": _texto(crudo["tipo_os"]),
                "lote": _identificador(crudo.get("lote")),
                "municipio": municipio,
                "barrio": barrio,
                "bkey": f"{municipio} | {barrio}" if municipio and barrio else None,
                "direccion": _texto(crudo["direccion"]),
                "cliente": _texto(crudo.get("cliente")),
                "tarifa": _texto(crudo.get("tarifa")),
                "deuda": _decimal(crudo.get("deuda")),
                "facturas": _entero(crudo.get("facturas")),
                "antiguedad": _entero(crudo.get("antiguedad")),
                "estado": _texto(crudo["estado"]),
                "tipo_suspension": _texto(crudo.get("tipo_suspension")),
                "estado_servicio": _texto(crudo.get("estado_servicio")),
                "comentario": _comentario(crudo.get("comentario")),
            }
        )

    _canonizar_comentarios(registros)

    logger.info(
        "Cargue: cabecera en la fila %s, %s filas leídas, %s órdenes con técnico (%s).",
        inicio + 1,
        leidas,
        len(registros),
        ", ".join(f"{motivo}={n}" for motivo, n in descartadas.items() if n) or "sin descartes",
    )
    return Cargue(
        ordenes=[Orden(**registro) for registro in registros],
        fila_cabecera=inicio + 1,
        leidas=leidas,
        descartadas=descartadas,
    )
